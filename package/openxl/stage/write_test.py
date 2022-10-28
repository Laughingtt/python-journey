import openpyxl  # 导入模块

# wb=openpyxl.Workbook() #实例化对象
wb = openpyxl.load_workbook("lianxi.xlsx")
ws = wb.active  # 选择活跃的sheet页
ws2 = wb.create_sheet("hello")
ws3 = wb["Hello1"]
ws3.append([1, 2, 3])
ws.append([2, 2, 3])  # 将列表添加到excel中
ws['A7'] = 666

A = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
B = [2, 3, 6, 7, 8, 9, 1, 2]
if len(A) > len(B):
    num = A
else:
    num = B

ws.append(["A", "B"])
for i in range(len(num)):
    if num == A:
        try:
            ws.append([A[i], B[i]])
        except:
            ws.append([A[i], " "])
    else:
        try:
            ws.append([A[i], B[i]])
        except:
            ws.append([" ", B[i]])

ws.cell(2, 3).value = "田田"

d = ws["A3"]
print(d.value)
col = ws["B"]
row = ws[1]
for i in row:
    print(i.value)
for i in ws.iter_cols(min_row=1, min_col=1, max_row=3, max_col=3):
    for j in i:
        print(j.value, end=" ")
    print()

print("row", ws.max_row)
print("column", ws.max_column)
del wb['hello']
wb.save("lianxi.xlsx")  # 保存excel
